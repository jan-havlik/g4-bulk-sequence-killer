import threading
from openpyxl import Workbook, load_workbook
from sequence import Sequence
from alive_progress import alive_bar

from helpers.complement import get_complementary_sequence

# how many base pairs will be included for scole calcaulation
SEQ_WINDOW_OVERLAP = 12

def run() -> None:
    """Script core function"""
    
    wb = load_workbook("g4.xlsx")

    sheet1 = wb.worksheets[0]
    sheet2 = wb.worksheets[1]
    
    expected_sheet_names = ["G4 1.6 all - MUTATED (1.1)", "G4 1.6 all - MUTATED (0.8)", "G4 1.8 strong - MUTATED (1.1)", "G4 1.8 strong - MUTATED (0.8)"]
    actual_sheet_names = wb.get_sheet_names()
    
    for sheet_name in expected_sheet_names:
        if sheet_name in actual_sheet_names:
            to_remove = wb[sheet_name]
            wb.remove(to_remove)
    
        
    for sheet_number in range(2, len(wb.worksheets)):
        wb.remove(wb.worksheets[sheet_number])
  

    sheet3 = wb.copy_worksheet(sheet1)
    sheet3.title = expected_sheet_names[0]
    
    sheet4 = wb.copy_worksheet(sheet1)
    sheet4.title = expected_sheet_names[1]

    sheet5 = wb.copy_worksheet(sheet2)
    sheet5.title = expected_sheet_names[2]
    
    sheet6 = wb.copy_worksheet(sheet2)
    sheet6.title = expected_sheet_names[3]

    with alive_bar(sheet1.max_row - 2) as bar:
        for row in range(2, sheet1.max_row):
            whole_sequence = sheet1.cell(row, 10).value.upper()
            
            orig_whole = whole_sequence

            for i in range(0, 186):  # 210 - 25
                partial_seq = Sequence(whole_sequence[i:i+25], score_threshold=1.1)
                partial_seq.calculate_score()
                
                orig_seq = partial_seq.sequence
                orig_score = partial_seq.score

                while partial_seq.over_threshold():
                    partial_seq.mutate()
                
                whole_sequence = whole_sequence[:i] + partial_seq.sequence + whole_sequence[i+25:]

            sheet3.cell(row, 10, whole_sequence)
            bar()

    with alive_bar(sheet1.max_row - 2) as bar:
        for row in range(2, sheet1.max_row):
            whole_sequence = sheet1.cell(row, 10).value.upper()
            
            orig_whole = whole_sequence

            for i in range(0, 186):  # 210 - 25
                partial_seq = Sequence(whole_sequence[i:i+25], score_threshold=0.8)
                partial_seq.calculate_score()
                
                orig_seq = partial_seq.sequence
                orig_score = partial_seq.score

                while partial_seq.over_threshold():
                    partial_seq.mutate()
                
                whole_sequence = whole_sequence[:i] + partial_seq.sequence + whole_sequence[i+25:]

            sheet4.cell(row, 10, whole_sequence)
            bar()

    with alive_bar(sheet2.max_row - 2) as bar:
        for row in range(2, sheet2.max_row):
            whole_sequence = sheet2.cell(row, 9).value.upper()
            
            orig_whole = whole_sequence

            for i in range(129, 135):  # 130:160 with 25 nucleotide window
                partial_seq = Sequence(whole_sequence[i:i+25], score_threshold=1.1)
                partial_seq.calculate_score()
                
                orig_seq = partial_seq.sequence
                orig_score = partial_seq.score

                while partial_seq.over_threshold():
                    partial_seq.mutate()
                
                whole_sequence = whole_sequence[:i] + partial_seq.sequence + whole_sequence[i+25:]
            sheet5.cell(row, 9, whole_sequence)
            bar()
    
    with alive_bar(sheet2.max_row - 2) as bar:
        for row in range(2, sheet2.max_row):
            whole_sequence = sheet2.cell(row, 9).value.upper()
            
            orig_whole = whole_sequence

            for i in range(129, 135):  # 130:160 with 25 nucleotide window
                partial_seq = Sequence(whole_sequence[i:i+25], score_threshold=0.8)
                partial_seq.calculate_score()
                
                orig_seq = partial_seq.sequence
                orig_score = partial_seq.score

                while partial_seq.over_threshold():
                    partial_seq.mutate()
                
                whole_sequence = whole_sequence[:i] + partial_seq.sequence + whole_sequence[i+25:]

            sheet6.cell(row, 9, whole_sequence)
            bar()

    wb.save('g4.xlsx')
if __name__ == "__main__":
    run()
